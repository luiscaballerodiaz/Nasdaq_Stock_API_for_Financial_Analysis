import csv
import requests
import numpy
import sys
import time
from openpyxl import Workbook


def get_value(dictionary, value):
    try:
        return float(dictionary['quarterlyReports'][0][value])
    except ValueError:
        return 0


url = 'https://www.alphavantage.co/query?function=BALANCE_SHEET&symbol='
apikey = 'xxx'
ticker_file = 'nasdaq_tickers.csv'
stock_data = []
tickers = []
with open(ticker_file) as f:
    reader = csv.reader(f)
    header_row = next(reader)
    for row in reader:
        if 'Banks' in row[-1] and '^' not in row[0] and float(row[5]) > 0 and 'United States' in row[6]:
            tickers.append(row[0])
            stock_data.append({
                'ticker': row[0],
                'market cap': round(float(row[5]) / 1e6, 2),
                'country': row[6]
            })

write_excel = numpy.zeros([len(stock_data), 12])
for stock, index in zip(stock_data, range(len(stock_data))):
    url_complete = url + stock['ticker'] + '&apikey=' + apikey
    r = requests.get(url_complete, verify=False)
    print("Status code:", r.status_code)
    response_dict = r.json()
    total_assets = get_value(response_dict, 'totalAssets')
    cash = get_value(response_dict, 'cashAndCashEquivalentsAtCarryingValue')
    inventory = get_value(response_dict, 'inventory')
    intangible_assets = get_value(response_dict, 'intangibleAssets')
    other_nocurrent_assets = get_value(response_dict, 'otherNonCurrentAssets')
    other_current_assets = get_value(response_dict, 'otherCurrentAssets')
    ppe = get_value(response_dict, 'propertyPlantEquipment')
    total_liabilities = get_value(response_dict, 'totalLiabilities')
    lt_debt = get_value(response_dict, 'longTermDebt')
    st_debt = get_value(response_dict, 'shortTermDebt')
    payables = get_value(response_dict, 'currentAccountsPayable')
    receivables = get_value(response_dict, 'currentNetReceivables')
    equity = get_value(response_dict, 'totalShareholderEquity')
    retained_earnings = get_value(response_dict, 'retainedEarnings')

    write_excel[index, 0] = round(total_assets / 1e6, 2)
    write_excel[index, 1] = round(cash / 1e6, 2)
    write_excel[index, 2] = round((total_assets - cash - intangible_assets - inventory - receivables
                                   - other_current_assets - other_nocurrent_assets) / 1e6, 2)
    write_excel[index, 3] = round(total_liabilities / 1e6, 2)
    write_excel[index, 4] = round((st_debt + lt_debt) / 1e6, 2)
    write_excel[index, 5] = round(equity / 1e6, 2)
    write_excel[index, 6] = round(retained_earnings / 1e6, 2)
    write_excel[index, 7] = round((total_liabilities - lt_debt - st_debt - payables) / 1e6, 2)
    try:
        write_excel[index, 8] = round(100 * write_excel[index, 2] / write_excel[index, 7], 2)
    except ZeroDivisionError:
        write_excel[index, 8] = 0
    try:
        write_excel[index, 9] = round(100 * write_excel[index, 5] / write_excel[index, 0], 2)
    except ZeroDivisionError:
        write_excel[index, 9] = 0
    try:
        write_excel[index, 10] = round(100 * write_excel[index, 1] / stock_data[index]['market cap'], 2)
    except ZeroDivisionError:
        write_excel[index, 10] = 0
    try:
        write_excel[index, 11] = round(100 * write_excel[index, 4] /
                                       (write_excel[index, 4] + write_excel[index, 5]), 2)
    except ZeroDivisionError:
        write_excel[index, 11] = 0

    time.sleep(12)

sheets = []
wb = Workbook()
sheets.append(wb.active)
sheets[0].title = 'NASDAQ BANK STOCKS'
# Define column width
column_number = range(1, 16)
for column in column_number:
    column = str(chr(64 + column))
    for sheet in sheets:
        sheet.column_dimensions[column].width = 22
for i in range(write_excel.shape[1]):  # Write data in excel sheet
    for j in range(write_excel.shape[0]):
        sheets[0].cell(j + 2, i + 4).value = round(write_excel[j, i], 2)  # Write data in a 2D loop
headers = ['ticker', 'market cap', 'country', 'total assets', 'cash', 'loans', 'liabilities', 'debt', 'equity',
           'retained earnings', 'deposits', 'loans to deposits', 'equity to assets', 'cash to market cap',
           'debt to total capitalization']
for header, index in zip(headers, range(len(headers))):
    sheets[0].cell(1, 1 + index).value = header
for stock, index in zip(stock_data, range(len(stock_data))):
    sheets[0].cell(2 + index, 1).value = stock['ticker']
    sheets[0].cell(2 + index, 2).value = stock['market cap']
    sheets[0].cell(2 + index, 3).value = stock['country']
try:
    wb.save('Nasdaq Stocks Assessment.xlsx')
except PermissionError:
    sys.exit('ERROR: Excel file open. Please close it to be modified')
